import copy
import numpy
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from settings import *
import pandas as pd
import os
import ast
import json


def load_data(
        path=settings.TEST_DATA_PATH,
        tables=settings.DF_NAMES,
        file_type='xlsx',
):
    """
    :param path: path to either SQL database or folder with tables of type CSV or Excel
    :param tables: list of files/tables
    :param file_type: type of file to read, can be xlsx (Excel file), CSV (comma-separated values), defaults to xlsx
    :return: list of pandas dataframes
    """

    if not settings.DEBUG and path==settings.TEST_DATA_PATH:
        raise FileNotFoundError(f"Path {settings.TEST_DATA_PATH} is reserved for debugging purposes, please specify a different path when using this function in release mode.")

    dataframes = {}
    for file in tables:
        if file_type == 'xlsx':
            try:
                dataframes[file] = pd.read_excel(os.path.join(path, file + '.' + file_type))
            except FileNotFoundError:
                dataframes[file] = pd.read_excel(os.path.join(path, file + '.' + 'ods'), engine="odf")
        if file_type == 'csv':
            dataframes[file] = pd.read_csv(os.path.join(path, file + '.' + file_type))

    dataframes['SSG_SUBJECTS']['teachers_ID'] = dataframes['SSG_SUBJECTS']['teachers_ID'].apply(ast.literal_eval)
    dataframes['SSG_SUBJECTS']['classroom_types'] = dataframes['SSG_SUBJECTS']['classroom_types'].apply(ast.literal_eval)

    dataframes['SSG_TEACHERS']['start_hour_index'] = dataframes['SSG_TEACHERS']['start_hour_index'].apply(ast.literal_eval)
    dataframes['SSG_TEACHERS']['end_hour_index'] = dataframes['SSG_TEACHERS']['end_hour_index'].apply(ast.literal_eval)
    dataframes['SSG_TEACHERS']['days'] = dataframes['SSG_TEACHERS']['days'].apply(ast.literal_eval)

    return list(dataframes.values())


def class_to_json(obj):
    def todict(_obj):
        if isinstance(_obj, list):
            _list = []
            for _value in _obj:
                _list.append(todict(_value))

            return _list
        elif (
            isinstance(_obj, (int, bool, float, str)) or
            _obj is None
        ):
            return _obj
        elif isinstance(_obj, numpy.int64):
            return int(_obj)
        else:
            _dict = {}

            iter_dict = _obj.items() if isinstance(_obj, dict) else _obj.__dict__.items()

            for key, value in iter_dict:
                if isinstance(key, numpy.int64):
                    key = int(key)

                if isinstance(value, list):
                    _dict[key] = []
                    for _value in value:
                        _dict[key].append(todict(_value))
                elif "class" in str(type(value)):
                    _dict[key] = todict(value)
                else:
                    _dict[key] = str(value)

            return _dict

    return todict(obj)


def schedule_to_json(schedule, file_path):
    schedule_dict = class_to_json(schedule)
    with open(f'{file_path}.json', 'a') as file:
        file.write('')
        json.dump(schedule_dict, file)

    return schedule_dict


def schedule_to_excel(schedule_dict, data, info, file_path):
    [
        lesson_hours_df,
        subject_names_df,
        subjects_df,
        teachers_df,
        classes_df,
        classrooms_df,
        classroom_types_df
    ] = copy.deepcopy(data)

    columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    try:
        wb = load_workbook(f'{file_path}.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
        ws.title = info['Title']

        ws.merge_cells('A1:E1')  # Merge cells for the title
        title_cell = ws['A1']
        title_cell.value = info['Title']
        title_cell.font = Font(size=20, bold=True)

        for i, key in enumerate(info):
            if key == 'Title':
                pass
            ws[f'A{i + 2}'] = key
            ws.merge_cells(f'B{i + 2}:E{i + 2}')
            value_cell = ws[f'B{i + 2}']
            value_cell.value = info[key]

    for class_name in schedule_dict.keys():
        ws = wb.create_sheet(title=f'Class_{class_name}')

        ws['A1'] = 'day/lesson'

        for lesson in range(len(lesson_hours_df)):
            ws[f'A{lesson + 2}'] = f"{lesson+1}."

        for i, day in enumerate(schedule_dict[class_name].keys()):
            ws[f'{columns[i + 1]}{1}'] = day

            for j, hour in enumerate(schedule_dict[class_name][day]):
                message = ''
                for subject in schedule_dict[class_name][day][j]:
                    try:
                        subject_name = subject_names_df.loc[
                            subject_names_df['subject_name_ID'] == subject['subject_id'], 'name'
                        ].values[0]
                    except IndexError:
                        subject_name = '---'

                    try:
                        teacher_name = " ".join(
                            teachers_df.loc[
                                teachers_df['teacher_ID'] == subject['teachers_id'][0], ['name', 'surname']
                            ].values[0]
                        )
                    except IndexError:
                        teacher_name = '---'

                    if subject['number_of_groups'] > 1:
                        message += f"gr.{subject['group']} : {subject_name} : {teacher_name}\n"
                    else:
                        message += f"{subject_name} : {teacher_name}"

                ws[f'{columns[i + 1]}{j + 2}'] = message

    wb.save(f'{file_path}.xlsx')
