import datetime as dt
import os
from datetime import datetime, timedelta
import pandas as pd
from django.conf import settings
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.core.files.storage import FileSystemStorage
from django.shortcuts import redirect
from django.shortcuts import render
from django.urls import reverse_lazy
from django.urls import reverse
from django.views.generic import FormView, TemplateView, View
from django.http import FileResponse, HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from ..forms import *
from ..schoolSchedule.generate import generate_schedule
from ..schoolSchedule.load_data import load_data, schedule_to_json
import urllib.parse


def update_context(request, kwargs, context):
    context['schedule_name'] = kwargs.get('schedule_name')
    context['model_name'] = kwargs.get('model')
    sort_by = request.GET.get('sort_by')
    if sort_by in [field.name for field in ScheduleList._meta.get_fields()]:
        context['schedule_list'] = ScheduleList.objects.filter(user_id=request.user).order_by(sort_by)
    else:
        context['schedule_list'] = ScheduleList.objects.filter(user_id=request.user)
    context['labels'] = [
        'lesson_hours',
        'classroom_types',
        'classrooms',
        'subject_names',
        'teachers',
        'classes',
        'subjects'
    ]
    return context


class SchedulesListView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = ScheduleListForm
    template_name = 'generatorApp/schedules.html'
    success_url = reverse_lazy('generatorApp:schedules_base')

    def form_valid(self, form):
        name = form.cleaned_data.get('name')
        description = form.cleaned_data.get('description')

        username = self.request.user.username
        self.success_url += f'{username}/{name}'

        if ScheduleList.objects.filter(user_id=self.request.user, name=name).exists():
            raise ValidationError('Schedule with this name exist')

        schedule = ScheduleList.objects.create(
            user_id=self.request.user,
            name=name,
            description=description,
            content=''
        )
        schedule.save()
        return redirect(self.success_url)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        return context


class ScheduleView(LoginRequiredMixin, TemplateView):
    template_name = 'generatorApp/schedule.html'

    def get_context_data(self, **kwargs):
        def schedule_str(schedule):
            if not schedule.content:
                return False

            schedule_content = json.loads(schedule.content)
            for class_id, days in schedule_content.items():
                for day, subjects in days.items():
                    for subject_list in subjects:
                        for subject in subject_list:
                            subject_name = schedule.subjectnames_set.filter(in_id=subject['subject_name_id']).first()
                            subject['subject_name_id'] = subject_name.name if subject_name else '---'

                            teacher_names = []
                            for teacher in subject['teachers_id']:
                                teacher = schedule.teachers_set.filter(in_id=teacher).first()
                                teacher_names.append(f'{teacher.name} {teacher.surname}' if teacher else '--- ---')
                            subject['teachers_id'] = teacher_names[-1]

                            classroom = schedule.classrooms_set.filter(in_id=subject['classroom_id']).first()
                            subject['classroom_id'] = classroom.name if classroom else '---'

                            lesson_hour = schedule.lessonhours_set.filter(in_id=subject['lesson_hour_id']).first()
                            if lesson_hour:
                                start_hour = datetime.strptime(lesson_hour.start_hour, '%H:%M:%S')
                                end_hour = (start_hour + dt.timedelta(minutes=45))

                                subject['lesson_hour_id'] = f'{start_hour.strftime("%H:%M")}-{end_hour.strftime("%H:%M")}'
                            else:
                                subject['lesson_hour_id'] = '---'

            return schedule_content

        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['schedule'] = schedule

        schedule_str = schedule_str(schedule)
        context['schedule_content'] = schedule_str if schedule_str else "Please import data!"
        context['classes'] = Classes.objects.filter(schedule_id=schedule)

        return context


class GenerateScheduleView(LoginRequiredMixin, View):
    login_url = reverse_lazy('generatorApp:login')
    success_url_name = 'generatorApp:schedule'

    def get(self, request, *args, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)

        username = context.get('username', request.user.username)
        schedule_name = context.get('schedule_name', 'schedule')    # TODO: wywalić error zamiast dawać default name jeśli nie znajdzie w context

        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        data = load_data(dtype='sql', schedule=schedule)

        schedule_settings = ScheduleSettings.objects.get(schedule_id=schedule)
        schedule_content = generate_schedule(data, json.loads(schedule_settings.content), log_file_name=schedule.id)

        if schedule_content:
            schedule.content = schedule_to_json(schedule_content)
            schedule.save()

            # Clear the error message after generating the schedule
            if 'warning_msg' in self.request.session:
                del self.request.session['warning_msg']

            success_url = reverse(self.success_url_name, kwargs={'username': username, 'schedule_name': schedule_name})
            return redirect(success_url)
        else:
            # TODO: Zmienić na błąd
            success_url = reverse(self.success_url_name, kwargs={'username': username, 'schedule_name': schedule_name})
            return redirect(success_url)


class LessonHoursView(LoginRequiredMixin, TemplateView, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = LessonHoursForm
    template_name = 'generatorApp/forms/lesson_hours.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = LessonHours.objects.filter(schedule_id=schedule).all()
        return context

    def form_valid(self, form, **kwargs):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        start_hour = form.cleaned_data.get('start_hour')

        last_obj = context['objects_list'].last()

        LessonHours.objects.create(
            in_id=int(last_obj.in_id)+1 if last_obj else 0,
            schedule_id=schedule,
            start_hour=start_hour,
            duration=45
        )
        return redirect(self.request.build_absolute_uri())


class ClassroomTypesView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = ClassroomTypesForm
    template_name = 'generatorApp/forms/classroom_types.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = ClassroomTypes.objects.filter(schedule_id=schedule).all()
        return context

    def form_valid(self, form, **kwargs):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        description = form.cleaned_data.get('description')

        last_obj = context['objects_list'].last()

        ClassroomTypes.objects.create(
            in_id=int(last_obj.in_id)+1 if last_obj else 0,
            schedule_id=schedule,
            description=description,
        )
        return redirect(self.request.build_absolute_uri())


class ClassroomsView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = ClassroomsForm
    template_name = 'generatorApp/forms/classrooms.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = Classrooms.objects.filter(schedule_id=schedule).all()
        context['queryset'] = ClassroomTypes.objects.filter(schedule_id=schedule)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        name = form.cleaned_data.get('name')
        type_id = self.request.POST.get('type-id')

        type_id = ClassroomTypes.objects.filter(schedule_id=schedule, in_id=type_id).first()

        last_obj = context['objects_list'].last()

        Classrooms.objects.create(
            in_id=int(last_obj.in_id) + 1 if last_obj else 0,
            schedule_id=schedule,
            type_id=type_id,
            name=name
        )
        return redirect(self.request.build_absolute_uri())


class TeachersView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = TeachersForm
    template_name = 'generatorApp/forms/teachers.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = Teachers.objects.filter(schedule_id=schedule).all()

        # listy elemntow do selektow
        context['classrooms_queryset'] = Classrooms.objects.filter(schedule_id=schedule)
        context['subjects_queryset'] = SubjectNames.objects.filter(schedule_id=schedule)
        context['lesson_hours_queryset'] = LessonHours.objects.filter(schedule_id=schedule)

        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']
        for row in context['objects_list']:
            row_days_list = json.loads(row.days)
            row.days = ["Yes" if row_days_list[i] else "No" for i, day in enumerate(days)]

        def add_duration(time_str, duration_minutes):
            time_format = "%H:%M:%S"
            time_obj = datetime.strptime(time_str, time_format)

            duration_obj = timedelta(minutes=duration_minutes)
            new_time_obj = time_obj + duration_obj
            new_time_str = new_time_obj.strftime("%H:%M")

            return new_time_str

        for row in context['objects_list']:
            row_start_hours_list = json.loads(row.start_hour_index)
            row.start_hour_index = [
                add_duration(
                    LessonHours.objects.filter(schedule_id=schedule, in_id=x).first().start_hour,
                    LessonHours.objects.filter(schedule_id=schedule, in_id=x).first().duration
                )
                for x in row_start_hours_list
            ]

        for row in context['objects_list']:
            row_end_hours_list = json.loads(row.end_hour_index)
            row.end_hour_index = [
                add_duration(
                    LessonHours.objects.filter(schedule_id=schedule, in_id=x).first().start_hour,
                    LessonHours.objects.filter(schedule_id=schedule, in_id=x).first().duration
                )
                for x in row_start_hours_list
            ]

        return context

    def form_valid(self, form):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        name = form.cleaned_data.get('name')
        surname = form.cleaned_data.get('surname')
        main_classroom_id = self.request.POST.get('main-classroom-id') \
            if self.request.POST.get('main-classroom-id') != "None" else None
    
        main_classroom_id = Classrooms.objects.filter(in_id=main_classroom_id, schedule_id=schedule).first()

        possible_subjects = [
            possible_subject if possible_subject != "None" else -1
            for possible_subject in self.request.POST.get('possible-subjects')
        ] if type(self.request.POST.getlist('possible-subjects')) == "<class 'list'>" \
            else self.request.POST.getlist('possible-subjects')

        days = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday']

        start_hours = []
        end_hours = []
        for day in days:
            start = LessonHours.objects.filter(
                schedule_id=schedule,
                start_hour=self.request.POST.get(f'start-hour-{day}')
            ).first()
            end = LessonHours.objects.filter(
                schedule_id=schedule,
                start_hour=self.request.POST.get(f'end-hour-{day}')
            ).first()
            start_hours.append(int(start.in_id))
            end_hours.append(int(end.in_id) if end else -1)

        possible_days = []
        for day in days:
            possible_days.append(1) if day in self.request.POST.getlist('days') else possible_days.append(0)

        last_obj = context['objects_list'].last()

        Teachers.objects.create(
            in_id=int(last_obj.in_id) + 1 if last_obj else 0,
            schedule_id=schedule,
            name=name,
            surname=surname,
            main_classroom_id=main_classroom_id,
            possible_subjects=json.dumps(', '.join(possible_subjects)),
            start_hour_index=json.dumps(start_hours),
            end_hour_index=json.dumps(end_hours),
            days=json.dumps(possible_days)
        )
        return redirect(self.request.build_absolute_uri())


class ClassesView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = ClassesForm
    template_name = 'generatorApp/forms/classes.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = Classes.objects.filter(schedule_id=schedule).all()

        # listy elemntow do selektow
        context['teachers_queryset'] = Teachers.objects.filter(schedule_id=schedule)
        context['lesson_hours_queryset'] = LessonHours.objects.filter(schedule_id=schedule)
        return context

    def form_valid(self, form):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        grade = form.cleaned_data.get('grade')
        class_signature = form.cleaned_data.get('class_signature')
        supervising_teacher = self.request.POST.get('supervising-teacher')
        supervising_teacher = Teachers.objects.filter(schedule_id=schedule, in_id=supervising_teacher).first()
        starting_lesson_hour = self.request.POST.get('starting-lesson-hour')
        starting_lesson_hour = LessonHours.objects.filter(schedule_id=schedule, in_id=starting_lesson_hour).first()

        last_obj = context['objects_list'].last()

        Classes.objects.create(
            in_id=int(last_obj.in_id) + 1 if last_obj else 0,
            schedule_id=schedule,
            supervising_teacher_id=supervising_teacher,
            starting_lesson_hour_id=starting_lesson_hour,
            grade=grade,
            class_signature=class_signature
        )
        return redirect(self.request.build_absolute_uri())


class SubjectNamesView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = SubjectNamesForm
    template_name = 'generatorApp/forms/subject_names.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = SubjectNames.objects.filter(schedule_id=schedule).all()
        return context

    def form_valid(self, form):
        # TODO: pozmieniac nazwy kolumn (posuuwac _id gdzie nie trza itp)
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        name = form.cleaned_data.get('name')

        last_obj = context['objects_list'].last()

        SubjectNames.objects.create(
            in_id=int(last_obj.in_id) + 1 if last_obj else 0,
            schedule_id=schedule,
            name=name
        )
        return redirect(self.request.build_absolute_uri())


class SubjectsView(LoginRequiredMixin, FormView):
    login_url = reverse_lazy('generatorApp:login')
    form_class = SubjectsForm
    template_name = 'generatorApp/forms/subjects.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = update_context(self.request, self.kwargs, context)
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        context['objects_list'] = Subject.objects.filter(schedule_id=schedule).all()

        # listy elemntow do selektow
        context['classes_queryset'] = Classes.objects.filter(schedule_id=schedule)
        context['subject_names_queryset'] = SubjectNames.objects.filter(schedule_id=schedule)
        context['teachers_queryset'] = Teachers.objects.filter(schedule_id=schedule)
        context['classroom_types_queryset'] = ClassroomTypes.objects.filter(schedule_id=schedule)

        for row in context['objects_list']:
            teachers_id = json.loads(row.teachers_id)
            row.teachers_id = [str(Teachers.objects.filter(schedule_id=schedule, in_id=x).first()) for x in teachers_id]

            classroom_types = json.loads(row.classroom_types)
            row.classroom_types = [str(ClassroomTypes.objects.filter(schedule_id=schedule, in_id=x).first()) for x in classroom_types]
        return context

    def form_valid(self, form):
        # TODO: walidacja zeby nie blo wiecej nauczycieli do przedmiotu niz jest grup
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        subject_count_in_week = form.cleaned_data.get('subject_count_in_week')
        number_of_groups = form.cleaned_data.get('number_of_groups')
        max_stack = form.cleaned_data.get('max_stack')

        class_id = self.request.POST.get('class-id')
        class_id = Classes.objects.filter(schedule_id=schedule, in_id=class_id).first()

        subject_name = self.request.POST.get('subject-name')
        subject_name = SubjectNames.objects.filter(schedule_id=schedule, in_id=subject_name).first()

        teachers = [int(x) for x in self.request.POST.getlist('teachers')]

        classroom_types = [int(x) for x in self.request.POST.getlist('classroom-type-id')]

        last_obj = context['objects_list'].last()

        Subject.objects.create(
            in_id=int(last_obj.in_id) + 1 if last_obj else 0,
            schedule_id=schedule,
            subject_count_in_week=subject_count_in_week,
            number_of_groups=number_of_groups,
            max_stack=max_stack,
            classes_id=class_id,
            subject_name_id=subject_name,
            teachers_id=teachers,
            classroom_types=classroom_types
        )
        return redirect(self.request.build_absolute_uri())


class DeleteScheduleView(LoginRequiredMixin, View):

    def get_context_data(self, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)
        return context

    def post(self, *args, **kwargs):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        ScheduleList.objects.get(user_id=self.request.user, name=schedule_name).delete()

        return redirect('/schedules/')


class DeleteDataView(LoginRequiredMixin, View):
    def get_context_data(self, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)

        # modele
        context['classes'] = Classes
        context['subject_names'] = SubjectNames
        context['teachers'] = Teachers
        context['classroom_types'] = ClassroomTypes
        context['classrooms'] = Classrooms
        context['subjects'] = Subject
        context['lesson_hours'] = LessonHours
        return context

    def post(self, *args, **kwargs):
        selected = self.request.POST.getlist('delete')
        context = self.get_context_data()
        model_name = context['model_name']
        model = context[model_name]
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        if selected:
            model.objects.filter(schedule_id=schedule, in_id__in=selected).delete()

        return redirect(self.request.META.get('HTTP_REFERER'))


class UploadDataView(LoginRequiredMixin, View):
    def get_context_data(self, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)
        return context

    def post(self, *args, **kwargs):
        context = self.get_context_data()
        model_name = context['model_name']
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        if self.request.FILES['file']:
            file = self.request.FILES['file']
            fs = FileSystemStorage()
            fs.save(file.name, file)

            allowed_extension = ['xlsx', 'ods']
            file_extension = file.name.split('.')[-1]

            if file_extension not in allowed_extension:
                fs.delete(file.name)
                # TODO: wyswietlic informacje o zlym rozszerzeniu
                return redirect(self.request.META.get('HTTP_REFERER'))

            file_path = os.path.join(settings.MEDIA_ROOT, file.name)

            if file_extension == 'ods':
                df = pd.read_excel(file_path, engine="odf")
            else:
                df = pd.read_excel(file_path)

            fs.delete(file.name)

            match model_name:
                case 'classes':
                    for index, row in df.iterrows():
                        in_id, grade, class_signature, supervising_teacher, starting_lesson_hour_id = row.values

                        if not Classes.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                supervising_teacher_id=Teachers.objects.get(_id=supervising_teacher, schedule_id=schedule),
                                starting_lesson_hour_id=LessonHours.objects.get(_id=starting_lesson_hour_id,
                                                                                schedule_id=schedule),
                                grade=grade,
                                class_signature=class_signature
                        ).exists():
                            Classes.objects.create(
                                in_id=in_id,
                                schedule_id=schedule,
                                supervising_teacher_id=Teachers.objects.get(_id=supervising_teacher, schedule_id=schedule),
                                starting_lesson_hour_id=LessonHours.objects.get(_id=starting_lesson_hour_id,
                                                                                schedule_id=schedule),
                                grade=grade,
                                class_signature=class_signature
                            )
                case 'classroom_types':
                    for index, row in df.iterrows():
                        in_id, description = row.values
                        if not ClassroomTypes.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                description=description
                        ).exists():
                            ClassroomTypes.objects.create(in_id=in_id, schedule_id=schedule, description=description)
                case 'classrooms':
                    for index, row in df.iterrows():
                        in_id, name, type_id = row.values
                        if not Classrooms.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                name=name,
                                type_id=ClassroomTypes.objects.get(id=type_id, schedule_id=schedule)
                        ).exists():
                            Classrooms.objects.create(
                                in_id=in_id, schedule_id=schedule,
                                name=name,
                                type_id=ClassroomTypes.objects.get(id=type_id)
                            )
                case 'lesson_hours':
                    for index, row in df.iterrows():
                        in_id, start_hour, duration = row.values
                        if not LessonHours.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                start_hour=start_hour,
                                duration=duration
                        ).exists():
                            LessonHours.objects.create(
                                in_id=in_id,
                                schedule_id=schedule,
                                start_hour=start_hour,
                                duration=duration
                            )
                case 'subject_names':
                    for index, row in df.iterrows():
                        in_id, name = row.values
                        if not SubjectNames.objects.filter(in_id=in_id, schedule_id=schedule, name=name).exists():
                            SubjectNames.objects.create(in_id=in_id, schedule_id=schedule, name=name)
                case 'teachers':
                    for index, row in df.iterrows():
                        in_id, name, surname, possible_subjects, start_hour_index, end_hour_index, days, main_classroom_id = row.values
                        if not Teachers.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                main_classroom_id=Classrooms.objects.get(_id=main_classroom_id,
                                                                         schedule_id=schedule) if main_classroom_id != 'Null' else None,
                                end_hour_index=end_hour_index,
                                days=days
                        ).exists():
                            Teachers.objects.create(
                                in_id=in_id,
                                schedule_id=schedule,
                                main_classroom_id=Classrooms.objects.get(_id=main_classroom_id,
                                                                         schedule_id=schedule) if main_classroom_id != 'Null' else None,
                                name=name,
                                surname=surname,
                                possible_subjects=possible_subjects,
                                start_hour_index=start_hour_index,
                                end_hour_index=end_hour_index,
                                days=days
                            )
                case 'subjects':
                    for index, row in df.iterrows():
                        (
                            in_id,
                            subject_name_id,
                            classes_id,
                            subject_count_in_week,
                            number_of_groups,
                            lesson_hour_id,
                            teachers_id,
                            classroom_id,
                            max_stack,
                            classroom_types
                        ) = row.values

                        if not Subject.objects.filter(
                                in_id=in_id,
                                schedule_id=schedule,
                                classes_id=Classes.objects.get(_id=classes_id, schedule_id=schedule),
                                subject_name_id=SubjectNames.objects.get(_id=subject_name_id, schedule_id=schedule),
                                lesson_hour_id=LessonHours.objects.get(_id=lesson_hour_id, schedule_id=schedule) if str(
                                    lesson_hour_id) != 'nan' else None,
                                teachers_id=teachers_id,
                                classroom_id=Classrooms.objects.get(_id=classroom_id, schedule_id=schedule) if str(
                                    classroom_id) != 'nan' else None,
                                subject_count_in_week=subject_count_in_week,
                                number_of_groups=number_of_groups,
                                max_stack=max_stack,
                                classroom_types=classroom_types
                        ).exists():
                            data = Subject(
                                in_id=in_id,
                                schedule_id=schedule,
                                classes_id=Classes.objects.get(_id=classes_id, schedule_id=schedule),
                                subject_name_id=SubjectNames.objects.get(_id=subject_name_id, schedule_id=schedule),
                                lesson_hour_id=LessonHours.objects.get(_id=lesson_hour_id, schedule_id=schedule) if str(
                                    lesson_hour_id) != 'nan' else None,
                                teachers_id=teachers_id,
                                classroom_id=Classrooms.objects.get(_id=classroom_id, schedule_id=schedule) if str(
                                    classroom_id) != 'nan' else None,
                                subject_count_in_week=subject_count_in_week,
                                number_of_groups=number_of_groups,
                                max_stack=max_stack,
                                classroom_types=classroom_types
                            )

                            if data.check_teachers(teachers_id, schedule):
                                data.save()
                # TODO: zwrocic ze podana nazwa nie pasuje do zadnego modelu/tabeli
                case _:
                    pass

        return redirect(self.request.META.get('HTTP_REFERER'))


class ScheduleSettingsView(LoginRequiredMixin, View):
    login_url = reverse_lazy('generatorApp:login')
    template_name = 'generatorApp/forms/settings.html'

    def get_context_data(self, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)

        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)
        settings = ScheduleSettings.objects.get(schedule_id=schedule)
        settings_dict = json.loads(settings.content)
        context['min_lessons'] = settings_dict['min_lessons_per_day']
        context['max_lessons'] = settings_dict['max_lessons_per_day']
        context['days'] = settings_dict['days']

        return context

    def get(self, *args, **kwargs):
        context = self.get_context_data()
        return render(self.request, self.template_name, context)

    def post(self, *args, **kwargs):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        min_lessons = self.request.POST.get('min-lessons')
        max_lessons = self.request.POST.get('max-lessons')
        days = self.request.POST.getlist('days')

        if int(min_lessons) >= int(max_lessons):
            context = self.get_context_data()
            context['error_msg'] = 'Min lessons per day must be lower than max lessons per day!!!'
            return render(self.request, self.template_name, context)

        settings = ScheduleSettings.objects.get(schedule_id=schedule)
        settings.content = json.dumps({
            "min_lessons_per_day": int(min_lessons),
            "max_lessons_per_day": int(max_lessons),
            "days": days,
        })
        settings.save()

        # Set the error message in the session
        self.request.session['warning_msg'] = 'Settings updated! Please regenerate your schedule.'

        return redirect(self.request.build_absolute_uri())


class ExportScheduleView(LoginRequiredMixin, View):
    login_url = reverse_lazy('generatorApp:login')
    template_name = 'generatorApp/forms/export_schedule.html'

    def get_context_data(self, **kwargs):
        context = {}
        context = update_context(self.request, self.kwargs, context)
        return context


    def schedule_to_excel(self, schedule, export_settings):
        schedule_dict = json.loads(schedule.content)
        file_path = os.path.join(settings.MEDIA_ROOT, str(schedule.id)+'.xlsx')

        columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

        try:
            wb = load_workbook(file_path)
        except FileNotFoundError:
            wb = Workbook()

        if 'Sheet' in wb.sheetnames:
            ws = wb['Sheet']
            ws.title = export_settings['Title']

            ws.merge_cells('A1:E1')  # Merge cells for the title
            title_cell = ws['A1']
            try:
                title_cell.value = export_settings['Title']
            except KeyError:
                title_cell.value = "School Schedule"
            title_cell.font = Font(size=20, bold=True)

            for i, key in enumerate(export_settings):
                if key != 'Title':
                    ws[f'A{i + 2}'] = key
                    ws.merge_cells(f'B{i + 2}:E{i + 2}')
                    value_cell = ws[f'B{i + 2}']
                    value_cell.value = export_settings[key]

        lesson_hours_df = LessonHours.objects.filter(schedule_id=schedule).values()
        classes_df = Classes.objects.filter(schedule_id=schedule).values()
        teachers_df = Teachers.objects.filter(schedule_id=schedule).values()
        subject_names_df = SubjectNames.objects.filter(schedule_id=schedule).values()
        classrooms_df = Classrooms.objects.filter(schedule_id=schedule).values()

        for class_name in schedule_dict.keys():
            ws = wb.create_sheet(title=f'Class_{class_name}')

            ws['A1'] = 'day/lesson'

            for lesson in range(len(lesson_hours_df)):
                ws[f'A{lesson + 2}'] = f"{lesson + 1}."

            class_info = classes_df.get(in_id=str(class_name))

            ws[f'A{len(lesson_hours_df) + 3}'] = f"Class:"
            try:
                ws[f'B{len(lesson_hours_df) + 3}'] = f"{class_info['grade']}{class_info['class_signature']}"
            except IndexError:
                ws[f'B{len(lesson_hours_df) + 3}'] = "---"

            supervising_teacher = teachers_df.get(in_id=str(class_info['supervising_teacher_id_id']))

            ws[f'D{len(lesson_hours_df) + 3}'] = f"Supervising teacher:"
            try:
                ws[f'E{len(lesson_hours_df) + 3}'] = f"{supervising_teacher['name']} {supervising_teacher['surname']}"
            except IndexError:
                ws[f'E{len(lesson_hours_df) + 3}'] = "---"

            for i, day in enumerate(schedule_dict[class_name].keys()):
                ws[f'{columns[i + 1]}{1}'] = day

                for j, hour in enumerate(schedule_dict[class_name][day]):
                    message = ''
                    for subject in schedule_dict[class_name][day][j]:
                        try:
                            print(subject.keys())
                            subject_name = subject_names_df.get(in_id=subject['subject_name_id'])
                        except SubjectNames.DoesNotExist:
                            subject_name = {'name': '---'}

                        try:
                            teacher_name = teachers_df.get(in_id=subject['teachers_id'][0])
                            teacher_name = f"{teacher_name['name']} {teacher_name['surname']}"
                        except Teachers.DoesNotExist:
                            teacher_name = '---'

                        try:
                            classroom_name = classrooms_df.get(in_id=subject['classroom_id'])['classroom_name']
                        except Classrooms.DoesNotExist:
                            classroom_name = '---'

                        if subject['number_of_groups'] > 1:
                            message += f"gr.{subject['group']} | {subject_name['name']} | {teacher_name} | class: {classroom_name}\n"
                        else:
                            message += f"{subject_name['name']} | {teacher_name} | class: {classroom_name}"

                    ws[f'{columns[i + 1]}{j + 2}'] = message

        wb.save(file_path)
        return file_path

    def get(self, *args, **kwargs):
        contex = self.get_context_data()
        return render(self.request, self.template_name, contex)

    def post(self, *args, **kwargs):
        context = self.get_context_data()
        schedule_name = context['schedule_name']
        schedule = ScheduleList.objects.get(user_id=self.request.user, name=schedule_name)

        file_title = self.request.POST.get('title', 'default_title')

        # Sanitize the file title to avoid issues with special characters
        sanitized_title = ''.join(char if char.isalnum() or char in (' ', '_', '-') else '_' for char in file_title)

        # Use a default name if file_title is empty
        if not sanitized_title:
            sanitized_title = "schedule"

        file_path = self.schedule_to_excel(
            schedule=schedule,
            export_settings={"Title": file_title}
        )

        with open(file_path, 'rb') as file:
            response = HttpResponse(
                file.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            # Use urllib.parse.quote to safely encode the filename for Content-Disposition
            encoded_file_title = urllib.parse.quote(sanitized_title)
            response['Content-Disposition'] = f'attachment; filename*=UTF-8\'\'{encoded_file_title}.xlsx'

        os.remove(file_path)

        return response
