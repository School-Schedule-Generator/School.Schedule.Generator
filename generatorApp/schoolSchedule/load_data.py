import copy
import sys
import numpy
import pandas as pd
import os
import ast
import json
import sqlite3
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from settings import *


def load_data(
        path=None,
        tables=settings.DF_NAMES,
        file_type='xlsx',
        schedule_id=None
):
    """
    :param path: path to either SQL database or folder with tables of type CSV or Excel
    :param tables: list of files/tables
    :param file_type: type of file to read, can be .xlsx/.ods (Excel file), .csv (comma-separated values) or sql, defaults to xlsx
    :param schedule_id: id of the schedule to pull from sql database
    :return: list of pandas dataframes or False if files don't match schedule data
    """

    if settings.DEBUG:
        path = settings.TEST_DATA_PATH
    elif not settings.DEBUG and path == settings.TEST_DATA_PATH:
        raise FileNotFoundError(f"Path {settings.TEST_DATA_PATH} is reserved for debugging purposes, please specify \
        a different path when using this function in release mode.")
    elif not settings.DEBUG and path is None:
        path = settings.BASE_DATA_PATH

    files_in_directory = os.listdir(path)

    if file_type == 'sql':
        try:
            conn = sqlite3.connect(settings.DATABASE_NAME)
        except sqlite3.Error:
            print(f'Failed to connect to database...', file=sys.stderr)
            return False

    dataframes = {}
    for table in tables:
        file_exists = False
        for file in files_in_directory:
            if file.startswith(table):
                file_exists = True

                if file_type == 'sql':
                    columns = settings.settings.SQLTABLES_NAMES[table]['Columns']
                    table_name = settings.SQLTABLES_NAMES[table]['Name']
                    sql_query = pd.read_sql_query(f'SELECT {", ".join(columns)} FROM {table_name} WHERE schedule_id={schedule_id}', conn)

                    dataframes[table] = pd.DataFrame(sql_query, columns=columns)
                elif file_type == 'xlsx':
                    try:
                        dataframes[table] = pd.read_excel(os.path.join(path, table + '.' + file_type))
                    except FileNotFoundError:
                        print(f'There is one or more files missing, please pass in: {file}.xlsx file or change directory of your\
                         data.', file=sys.stderr)
                        return False
                elif file_type == 'ods':
                    try:
                        dataframes[table] = pd.read_excel(os.path.join(path, table + '.' + 'ods'), engine="odf")
                    except FileNotFoundError:
                        print(f'There is one or more files missing, please pass in: {file}.ods file or change directory\
                         of your data.', file=sys.stderr)
                        return False
                elif file_type == 'csv':
                    try:
                        dataframes[table] = pd.read_csv(os.path.join(path, table + '.' + file_type))
                    except FileNotFoundError:
                        print(f'There is one or more files missing, please pass in: {file}.csv file or change directory\
                         of your data.', file=sys.stderr)
                        return False
                else:
                    print(f'File type: {file_type} is not suported, please pass in files in .xlsx, .ods or .csv',
                          file=sys.stderr)
                    return False

                if len(set(settings.COLUMN_NAMES[table])) < len(set(dataframes[table].columns.values)):
                    print(
                        f"There is too many columns in {table}\n"
                        f"Unwanted columns: {set(dataframes[table].columns.values) - set(settings.COLUMN_NAMES[table])}",
                        file=sys.stderr
                    )
                    return False

                elif set(settings.COLUMN_NAMES[table]) != set(dataframes[table].columns.values):
                    print(
                        f"Column names in passed in data file: \"{table}\" don't mach default schedule column names:\n"
                        f"Passed in: {set(dataframes[table].columns.values) - set(settings.COLUMN_NAMES[table])}\n"
                        f"Needed: {set(settings.COLUMN_NAMES[table]) - set(dataframes[table].columns.values)}",
                        file=sys.stderr
                    )
                    return False

        if not file_exists:
            print(f'There is one or more files missing, please pass in: {file} file or change directory of your data.', file=sys.stderr)
            return False

    dataframes['SSG_SUBJECTS']['teachers_id'] = dataframes['SSG_SUBJECTS']['teachers_id'].apply(ast.literal_eval)
    dataframes['SSG_SUBJECTS']['classroom_types'] = dataframes['SSG_SUBJECTS']['classroom_types'].apply(ast.literal_eval)

    dataframes['SSG_TEACHERS']['start_hour_index'] = dataframes['SSG_TEACHERS']['start_hour_index'].apply(ast.literal_eval)
    dataframes['SSG_TEACHERS']['end_hour_index'] = dataframes['SSG_TEACHERS']['end_hour_index'].apply(ast.literal_eval)
    dataframes['SSG_TEACHERS']['days'] = dataframes['SSG_TEACHERS']['days'].apply(ast.literal_eval)

    return list(dataframes.values())


def class_to_json(obj):
    def todict(_obj):
        if isinstance(_obj, list):
            _list = []
            for _value in _obj:
                _list.append(todict(_value))

            return _list
        elif (
            isinstance(_obj, (int, bool, float, str)) or
            _obj is None
        ):
            return _obj
        elif isinstance(_obj, numpy.int64):
            return int(_obj)
        else:
            _dict = {}

            iter_dict = _obj.items() if isinstance(_obj, dict) else _obj.__dict__.items()

            for key, value in iter_dict:
                if isinstance(key, numpy.int64):
                    key = int(key)

                if isinstance(value, list):
                    _dict[key] = []
                    for _value in value:
                        _dict[key].append(todict(_value))
                elif "class" in str(type(value)):
                    _dict[key] = todict(value)
                else:
                    _dict[key] = str(value)

            return _dict

    return todict(obj)


def schedule_to_json(schedule, file_path):
    schedule_dict = class_to_json(schedule)
    with open(f'{file_path}.json', 'a') as file:
        file.write('')
        json.dump(schedule_dict, file)

    return schedule_dict


def schedule_to_excel(schedule_dict, data, info, file_path):
    [
        lesson_hours_df,
        subject_names_df,
        _,
        teachers_df,
        classes_df,
        classrooms_df,
        _
    ] = copy.deepcopy(data)

    columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    try:
        wb = load_workbook(f'{file_path}.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    if 'Sheet' in wb.sheetnames:
        ws = wb['Sheet']
        ws.title = info['Title']

        ws.merge_cells('A1:E1')  # Merge cells for the title
        title_cell = ws['A1']
        try:
            title_cell.value = info['Title']
        except KeyError:
            title_cell.value = "School Schedule"
        title_cell.font = Font(size=20, bold=True)

        for i, key in enumerate(info):
            if key == 'Title':
                pass
            ws[f'A{i + 2}'] = key
            ws.merge_cells(f'B{i + 2}:E{i + 2}')
            value_cell = ws[f'B{i + 2}']
            value_cell.value = info[key]

    for class_name in schedule_dict.keys():
        ws = wb.create_sheet(title=f'Class_{class_name}')

        ws['A1'] = 'day/lesson'

        for lesson in range(len(lesson_hours_df)):
            ws[f'A{lesson + 2}'] = f"{lesson+1}."

        class_info = classes_df.loc[classes_df['Class_id'] == class_name]
        ws[f'A{len(lesson_hours_df) + 3}'] = f"Class:"
        try:
            ws[f'B{len(lesson_hours_df) + 3}'] =  f"{class_info['grade'].values[0]}{class_info['class_signature'].values[0]}"
        except IndexError:
            ws[f'B{len(lesson_hours_df) + 3}'] = "---"

        supervising_teacher = teachers_df.loc[teachers_df['teacher_id'] == class_info['supervising_teacher'].values[0], ['name', 'surname']].values[0]
        ws[f'D{len(lesson_hours_df) + 3}'] = f"Supervising teacher:"
        try:
            ws[f'E{len(lesson_hours_df) + 3}'] = f"{supervising_teacher[0]} {supervising_teacher[1]}"
        except IndexError:
            ws[f'E{len(lesson_hours_df) + 3}'] = "---"


        for i, day in enumerate(schedule_dict[class_name].keys()):
            ws[f'{columns[i + 1]}{1}'] = day

            for j, hour in enumerate(schedule_dict[class_name][day]):
                message = ''
                for subject in schedule_dict[class_name][day][j]:
                    try:
                        subject_name = subject_names_df.loc[
                            subject_names_df['subject_name_id'] == subject['subject_id'], 'name'
                        ].values[0]
                    except IndexError:
                        subject_name = '---'

                    try:
                        teacher_name = " ".join(
                            teachers_df.loc[
                                teachers_df['teacher_id'] == subject['teachers_id'][0], ['name', 'surname']
                            ].values[0]
                        )
                    except IndexError:
                        teacher_name = '---'

                    try:
                        classroom_name = classrooms_df.loc[
                            classrooms_df['classroom_id'] == subject['classroom_id'], 'classroom_name'
                        ].values[0]
                    except IndexError:
                        teacher_name = '---'

                    if subject['number_of_groups'] > 1:
                        message += f"gr.{subject['group']} | {subject_name} | {teacher_name} | class: {classroom_name}\n"
                    else:
                        message += f"{subject_name} | {teacher_name} | class: {classroom_name}"

                ws[f'{columns[i + 1]}{j + 2}'] = message

    wb.save(f'{file_path}.xlsx')
